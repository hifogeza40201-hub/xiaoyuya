# 餐饮自动化报表系统

## 系统架构

```
┌─────────────────────────────────────────────────────────────────┐
│                     自动化报表系统                                │
├─────────────────────────────────────────────────────────────────┤
│  ┌──────────┐  ┌──────────┐  ┌──────────┐  ┌──────────┐        │
│  │ 数据采集  │→│ 数据处理  │→│ 报表生成  │→│ 分发推送  │        │
│  └──────────┘  └──────────┘  └──────────┘  └──────────┘        │
├─────────────────────────────────────────────────────────────────┤
│  调度器：APScheduler / Cron                                    │
│  存储：MySQL + Parquet                                          │
│  报告：Excel / PDF / 邮件 / 钉钉                                │
└─────────────────────────────────────────────────────────────────┘
```

---

## 完整代码实现

```python
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
餐饮自动化报表系统
支持：日报、周报、月报自动生成与推送
"""

import os
import json
import logging
from datetime import datetime, timedelta
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple
from pathlib import Path

import pandas as pd
import numpy as np
from sqlalchemy import create_engine
import matplotlib.pyplot as plt
import seaborn as sns

# 报告生成
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

# 邮件发送
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

# 钉钉推送
import requests

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('report_system.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


# ============================================================================
# 配置类
# ============================================================================

@dataclass
class ReportConfig:
    """报表配置"""
    # 数据库配置
    db_host: str = 'localhost'
    db_port: int = 3306
    db_user: str = 'root'
    db_password: str = 'password'
    db_name: str = 'restaurant_db'
    
    # 邮件配置
    smtp_host: str = 'smtp.example.com'
    smtp_port: int = 587
    smtp_user: str = 'report@example.com'
    smtp_password: str = 'email_password'
    email_recipients: List[str] = None
    
    # 钉钉配置
    dingtalk_webhook: str = 'https://oapi.dingtalk.com/robot/send?access_token=xxx'
    dingtalk_secret: str = 'secret'
    
    # 报表配置
    output_dir: str = './reports'
    logo_path: Optional[str] = None
    
    def __post_init__(self):
        if self.email_recipients is None:
            self.email_recipients = ['manager@example.com']
        os.makedirs(self.output_dir, exist_ok=True)


# ============================================================================
# 数据层
# ============================================================================

class DataCollector:
    """数据采集器"""
    
    def __init__(self, config: ReportConfig):
        self.config = config
        self.engine = create_engine(
            f'mysql+pymysql://{config.db_user}:{config.db_password}'
            f'@{config.db_host}:{config.db_port}/{config.db_name}'
        )
    
    def get_sales_summary(self, start_date: str, end_date: str) -> pd.DataFrame:
        """获取销售汇总数据"""
        query = f"""
        SELECT 
            DATE(o.created_at) as date,
            COUNT(DISTINCT o.order_id) as order_count,
            COUNT(DISTINCT o.user_id) as customer_count,
            SUM(o.pay_amount) as total_amount,
            AVG(o.pay_amount) as avg_order_value,
            SUM(o.discount_amount) as discount_amount
        FROM orders o
        WHERE o.created_at BETWEEN '{start_date}' AND '{end_date}'
          AND o.status = 'PAID'
        GROUP BY DATE(o.created_at)
        ORDER BY date
        """
        return pd.read_sql(query, self.engine)
    
    def get_category_sales(self, start_date: str, end_date: str) -> pd.DataFrame:
        """获取品类销售数据"""
        query = f"""
        SELECT 
            c.category_name,
            SUM(od.quantity) as total_qty,
            SUM(od.total_amount) as total_amount,
            COUNT(DISTINCT od.order_id) as order_count
        FROM order_details od
        JOIN orders o ON od.order_id = o.order_id
        JOIN categories c ON od.category_id = c.category_id
        WHERE o.created_at BETWEEN '{start_date}' AND '{end_date}'
          AND o.status = 'PAID'
        GROUP BY c.category_id, c.category_name
        ORDER BY total_amount DESC
        """
        return pd.read_sql(query, self.engine)
    
    def get_top_products(self, start_date: str, end_date: str, top_n: int = 20) -> pd.DataFrame:
        """获取热销商品"""
        query = f"""
        SELECT 
            od.sku_name,
            SUM(od.quantity) as total_qty,
            SUM(od.total_amount) as total_amount,
            AVG(od.unit_price) as avg_price
        FROM order_details od
        JOIN orders o ON od.order_id = o.order_id
        WHERE o.created_at BETWEEN '{start_date}' AND '{end_date}'
          AND o.status = 'PAID'
        GROUP BY od.sku_id, od.sku_name
        ORDER BY total_amount DESC
        LIMIT {top_n}
        """
        return pd.read_sql(query, self.engine)
    
    def get_hourly_distribution(self, start_date: str, end_date: str) -> pd.DataFrame:
        """获取时段分布"""
        query = f"""
        SELECT 
            HOUR(o.created_at) as hour,
            COUNT(*) as order_count,
            SUM(o.pay_amount) as total_amount
        FROM orders o
        WHERE o.created_at BETWEEN '{start_date}' AND '{end_date}'
          AND o.status = 'PAID'
        GROUP BY HOUR(o.created_at)
        ORDER BY hour
        """
        return pd.read_sql(query, self.engine)
    
    def get_payment_methods(self, start_date: str, end_date: str) -> pd.DataFrame:
        """获取支付方式分布"""
        query = f"""
        SELECT 
            pay_method,
            COUNT(*) as order_count,
            SUM(pay_amount) as total_amount,
            AVG(pay_amount) as avg_amount
        FROM orders
        WHERE created_at BETWEEN '{start_date}' AND '{end_date}'
          AND status = 'PAID'
        GROUP BY pay_method
        ORDER BY total_amount DESC
        """
        return pd.read_sql(query, self.engine)
    
    def get_shop_comparison(self, start_date: str, end_date: str) -> pd.DataFrame:
        """获取门店对比数据"""
        query = f"""
        SELECT 
            s.shop_name,
            s.city,
            COUNT(DISTINCT o.order_id) as order_count,
            COUNT(DISTINCT o.user_id) as customer_count,
            SUM(o.pay_amount) as total_amount,
            AVG(o.pay_amount) as avg_order_value
        FROM orders o
        JOIN shops s ON o.shop_id = s.shop_id
        WHERE o.created_at BETWEEN '{start_date}' AND '{end_date}'
          AND o.status = 'PAID'
        GROUP BY s.shop_id, s.shop_name, s.city
        ORDER BY total_amount DESC
        """
        return pd.read_sql(query, self.engine)


# ============================================================================
# 报表生成器
# ============================================================================

class ReportGenerator:
    """报表生成器"""
    
    def __init__(self, config: ReportConfig):
        self.config = config
        self.collector = DataCollector(config)
        plt.rcParams['font.sans-serif'] = ['SimHei', 'DejaVu Sans']
        plt.rcParams['axes.unicode_minus'] = False
    
    def generate_daily_report(self, report_date: datetime = None) -> str:
        """生成日报"""
        if report_date is None:
            report_date = datetime.now() - timedelta(days=1)
        
        date_str = report_date.strftime('%Y-%m-%d')
        start_date = date_str + ' 00:00:00'
        end_date = date_str + ' 23:59:59'
        
        logger.info(f"Generating daily report for {date_str}")
        
        # 采集数据
        data = {
            'summary': self.collector.get_sales_summary(start_date, end_date),
            'category': self.collector.get_category_sales(start_date, end_date),
            'top_products': self.collector.get_top_products(start_date, end_date, top_n=15),
            'hourly': self.collector.get_hourly_distribution(start_date, end_date),
            'payment': self.collector.get_payment_methods(start_date, end_date),
            'shops': self.collector.get_shop_comparison(start_date, end_date)
        }
        
        # 生成图表
        charts = self._create_charts(data, date_str)
        
        # 生成Excel报告
        report_path = self._create_excel_report(data, charts, date_str, 'daily')
        
        logger.info(f"Daily report generated: {report_path}")
        return report_path
    
    def generate_weekly_report(self, end_date: datetime = None) -> str:
        """生成周报"""
        if end_date is None:
            end_date = datetime.now() - timedelta(days=1)
        
        start_date = end_date - timedelta(days=6)
        start_str = start_date.strftime('%Y-%m-%d')
        end_str = end_date.strftime('%Y-%m-%d')
        
        logger.info(f"Generating weekly report from {start_str} to {end_str}")
        
        start_dt = start_str + ' 00:00:00'
        end_dt = end_str + ' 23:59:59'
        
        data = {
            'summary': self.collector.get_sales_summary(start_dt, end_dt),
            'category': self.collector.get_category_sales(start_dt, end_dt),
            'top_products': self.collector.get_top_products(start_dt, end_dt, top_n=20),
            'hourly': self.collector.get_hourly_distribution(start_dt, end_dt),
            'payment': self.collector.get_payment_methods(start_dt, end_dt),
            'shops': self.collector.get_shop_comparison(start_dt, end_dt)
        }
        
        charts = self._create_charts(data, f"{start_str}_to_{end_str}")
        report_path = self._create_excel_report(data, charts, f"{start_str}_to_{end_str}", 'weekly')
        
        logger.info(f"Weekly report generated: {report_path}")
        return report_path
    
    def _create_charts(self, data: Dict, period: str) -> Dict[str, str]:
        """创建可视化图表"""
        charts = {}
        chart_dir = os.path.join(self.config.output_dir, 'charts')
        os.makedirs(chart_dir, exist_ok=True)
        
        # 1. 销售趋势图
        if not data['summary'].empty:
            fig, ax = plt.subplots(figsize=(12, 5))
            ax.plot(data['summary']['date'], data['summary']['total_amount'], 
                   marker='o', linewidth=2, markersize=4)
            ax.set_title(f'Sales Trend - {period}', fontsize=14)
            ax.set_xlabel('Date')
            ax.set_ylabel('Amount')
            ax.grid(True, alpha=0.3)
            plt.xticks(rotation=45)
            plt.tight_layout()
            charts['trend'] = os.path.join(chart_dir, f'trend_{period}.png')
            fig.savefig(charts['trend'], dpi=150, bbox_inches='tight')
            plt.close()
        
        # 2. 品类占比饼图
        if not data['category'].empty:
            fig, ax = plt.subplots(figsize=(8, 8))
            top_cats = data['category'].head(8)
            others = data['category'].iloc[8:]['total_amount'].sum()
            if others > 0:
                top_cats = pd.concat([
                    top_cats,
                    pd.DataFrame([{'category_name': 'Others', 'total_amount': others}])
                ])
            ax.pie(top_cats['total_amount'], labels=top_cats['category_name'], 
                  autopct='%1.1f%%', startangle=90)
            ax.set_title(f'Category Distribution - {period}', fontsize=14)
            charts['category'] = os.path.join(chart_dir, f'category_{period}.png')
            fig.savefig(charts['category'], dpi=150, bbox_inches='tight')
            plt.close()
        
        # 3. 时段分布图
        if not data['hourly'].empty:
            fig, ax = plt.subplots(figsize=(12, 5))
            ax.bar(data['hourly']['hour'], data['hourly']['order_count'], 
                  color='steelblue', alpha=0.8)
            ax.set_title(f'Hourly Distribution - {period}', fontsize=14)
            ax.set_xlabel('Hour')
            ax.set_ylabel('Order Count')
            ax.set_xticks(range(0, 24))
            plt.tight_layout()
            charts['hourly'] = os.path.join(chart_dir, f'hourly_{period}.png')
            fig.savefig(charts['hourly'], dpi=150, bbox_inches='tight')
            plt.close()
        
        return charts
    
    def _create_excel_report(self, data: Dict, charts: Dict, 
                            period: str, report_type: str) -> str:
        """创建Excel报表"""
        
        # 计算关键指标
        summary = data['summary']
        total_orders = summary['order_count'].sum() if not summary.empty else 0
        total_amount = summary['total_amount'].sum() if not summary.empty else 0
        total_customers = summary['customer_count'].sum() if not summary.empty else 0
        avg_order_value = total_amount / total_orders if total_orders > 0 else 0
        
        # 创建工作簿
        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "数据概览"
        
        # 样式定义
        title_font = Font(size=18, bold=True, color="FFFFFF")
        title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        metric_font = Font(size=24, bold=True, color="4472C4")
        label_font = Font(size=11, color="666666")
        
        # 标题
        ws_summary.merge_cells('A1:E1')
        ws_summary['A1'] = f"餐饮经营{'日报' if report_type == 'daily' else '周报'} - {period}"
        ws_summary['A1'].font = title_font
        ws_summary['A1'].fill = title_fill
        ws_summary['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws_summary.row_dimensions[1].height = 35
        
        # 关键指标卡片
        metrics = [
            ('B', '订单总数', f'{total_orders:,}'),
            ('D', '营业额', f'¥{total_amount:,.2f}'),
            ('F', '客单价', f'¥{avg_order_value:.2f}'),
            ('H', '顾客数', f'{total_customers:,}')
        ]
        
        for col, label, value in metrics:
            cell_label = ws_summary[f'{col}3']
            cell_value = ws_summary[f'{col}4']
            cell_label.value = label
            cell_label.font = label_font
            cell_label.alignment = Alignment(horizontal='center')
            cell_value.value = value
            cell_value.font = metric_font
            cell_value.alignment = Alignment(horizontal='center')
        
        # 品类销售表
        row = 7
        ws_summary[f'A{row}'] = '品类销售TOP10'
        ws_summary[f'A{row}'].font = Font(size=14, bold=True)
        
        row += 1
        headers = ['排名', '品类', '销售量', '销售额', '订单数']
        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row, col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        if not data['category'].empty:
            for idx, (_, row_data) in enumerate(data['category'].head(10).iterrows(), 1):
                row += 1
                ws_summary.cell(row, 1).value = idx
                ws_summary.cell(row, 2).value = row_data.get('category_name', '')
                ws_summary.cell(row, 3).value = row_data.get('total_qty', 0)
                ws_summary.cell(row, 4).value = row_data.get('total_amount', 0)
                ws_summary.cell(row, 5).value = row_data.get('order_count', 0)
        
        # 热销商品表
        row += 3
        ws_summary[f'A{row}'] = '热销商品TOP15'
        ws_summary[f'A{row}'].font = Font(size=14, bold=True)
        
        row += 1
        headers = ['排名', '商品名称', '销量', '销售额', '均价']
        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row, col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        if not data['top_products'].empty:
            for idx, (_, row_data) in enumerate(data['top_products'].iterrows(), 1):
                row += 1
                ws_summary.cell(row, 1).value = idx
                ws_summary.cell(row, 2).value = row_data.get('sku_name', '')
                ws_summary.cell(row, 3).value = row_data.get('total_qty', 0)
                ws_summary.cell(row, 4).value = row_data.get('total_amount', 0)
                ws_summary.cell(row, 5).value = row_data.get('avg_price', 0)
        
        # 门店对比表
        ws_shops = wb.create_sheet("门店对比")
        if not data['shops'].empty:
            for r_idx, row in enumerate(dataframe_to_rows(data['shops'], index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_shops.cell(r_idx, c_idx, value)
                    if r_idx == 1:
                        cell.font = header_font
                        cell.fill = header_fill
        
        # 时段分析表
        ws_hourly = wb.create_sheet("时段分析")
        if not data['hourly'].empty:
            for r_idx, row in enumerate(dataframe_to_rows(data['hourly'], index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_hourly.cell(r_idx, c_idx, value)
                    if r_idx == 1:
                        cell.font = header_font
                        cell.fill = header_fill
        
        # 原始数据表
        ws_raw = wb.create_sheet("原始数据")
        if not summary.empty:
            for r_idx, row in enumerate(dataframe_to_rows(summary, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_raw.cell(r_idx, c_idx, value)
                    if r_idx == 1:
                        cell.font = header_font
                        cell.fill = header_fill
        
        # 调整列宽
        for ws in [ws_summary, ws_shops, ws_hourly, ws_raw]:
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
        
        # 保存文件
        filename = f"{report_type}_report_{period.replace('-', '')}.xlsx"
        filepath = os.path.join(self.config.output_dir, filename)
        wb.save(filepath)
        
        return filepath


# ============================================================================
# 分发器
# ============================================================================

class ReportDistributor:
    """报表分发器"""
    
    def __init__(self, config: ReportConfig):
        self.config = config
    
    def send_email(self, report_path: str, subject: str = None, 
                   body: str = None) -> bool:
        """发送邮件报告"""
        try:
            if subject is None:
                subject = f"餐饮经营报表 - {os.path.basename(report_path)}"
            if body is None:
                body = "请查收附件中的经营报表。"
            
            msg = MIMEMultipart()
            msg['From'] = self.config.smtp_user
            msg['To'] = ', '.join(self.config.email_recipients)
            msg['Subject'] = subject
            
            msg.attach(MIMEText(body, 'plain', 'utf-8'))
            
            # 添加附件
            with open(report_path, 'rb') as f:
                attachment = MIMEApplication(f.read())
                attachment.add_header('Content-Disposition', 'attachment', 
                                    filename=os.path.basename(report_path))
                msg.attach(attachment)
            
            # 发送
            with smtplib.SMTP(self.config.smtp_host, self.config.smtp_port) as server:
                server.starttls()
                server.login(self.config.smtp_user, self.config.smtp_password)
                server.send_message(msg)
            
            logger.info(f"Email sent to {self.config.email_recipients}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to send email: {e}")
            return False
    
    def send_dingtalk(self, message: str, report_path: str = None) -> bool:
        """发送钉钉消息"""
        try:
            import hmac
            import hashlib
            import base64
            import urllib.parse
            import time
            
            # 计算签名
            timestamp = str(round(time.time() * 1000))
            secret_enc = self.config.dingtalk_secret.encode('utf-8')
            string_to_sign = f'{timestamp}\n{self.config.dingtalk_secret}'
            string_to_sign_enc = string_to_sign.encode('utf-8')
            hmac_code = hmac.new(secret_enc, string_to_sign_enc, 
                                digestmod=hashlib.sha256).digest()
            sign = urllib.parse.quote_plus(base64.b64encode(hmac_code))
            
            webhook = f"{self.config.dingtalk_webhook}&timestamp={timestamp}&sign={sign}"
            
            data = {
                "msgtype": "markdown",
                "markdown": {
                    "title": "经营报表",
                    "text": message
                }
            }
            
            response = requests.post(webhook, json=data, timeout=30)
            result = response.json()
            
            if result.get('errcode') == 0:
                logger.info("DingTalk message sent successfully")
                return True
            else:
                logger.error(f"DingTalk error: {result}")
                return False
                
        except Exception as e:
            logger.error(f"Failed to send DingTalk message: {e}")
            return False


# ============================================================================
# 调度器
# ============================================================================

class ReportScheduler:
    """报表调度器"""
    
    def __init__(self, config: ReportConfig):
        self.config = config
        self.generator = ReportGenerator(config)
        self.distributor = ReportDistributor(config)
    
    def run_daily_report(self):
        """执行日报任务"""
        try:
            report_path = self.generator.generate_daily_report()
            
            # 生成摘要
            summary = self._generate_summary(report_path, 'daily')
            
            # 发送邮件
            self.distributor.send_email(
                report_path,
                subject=f"餐饮经营日报 - {datetime.now().strftime('%Y-%m-%d')}",
                body=summary
            )
            
            # 钉钉推送
            markdown = f"""## 📊 餐饮经营日报
            
{summary}

[查看完整报表](file://{report_path})
            """
            self.distributor.send_dingtalk(markdown)
            
            logger.info("Daily report task completed")
            
        except Exception as e:
            logger.error(f"Daily report task failed: {e}")
    
    def run_weekly_report(self):
        """执行周报任务"""
        try:
            report_path = self.generator.generate_weekly_report()
            
            summary = self._generate_summary(report_path, 'weekly')
            
            self.distributor.send_email(
                report_path,
                subject=f"餐饮经营周报 - {datetime.now().strftime('%Y-%m-%d')}",
                body=summary
            )
            
            markdown = f"""## 📈 餐饮经营周报
            
{summary}

[查看完整报表](file://{report_path})
            """
            self.distributor.send_dingtalk(markdown)
            
            logger.info("Weekly report task completed")
            
        except Exception as e:
            logger.error(f"Weekly report task failed: {e}")
    
    def _generate_summary(self, report_path: str, report_type: str) -> str:
        """生成报告摘要"""
        # 这里可以读取Excel文件生成更详细的摘要
        summary = f"""
报表生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
报表文件：{os.path.basename(report_path)}

报表包含内容：
- 数据概览（关键指标、品类销售、热销商品）
- 门店对比分析
- 时段销售分布
- 原始明细数据

请查看附件获取完整报表。
        """
        return summary


# ============================================================================
# 主程序
# ============================================================================

def main():
    """主函数"""
    # 加载配置
    config = ReportConfig()
    
    # 如果有配置文件则加载
    config_file = 'report_config.json'
    if os.path.exists(config_file):
        with open(config_file, 'r', encoding='utf-8') as f:
            config_dict = json.load(f)
            for key, value in config_dict.items():
                if hasattr(config, key):
                    setattr(config, key, value)
    
    # 创建调度器
    scheduler = ReportScheduler(config)
    
    # 命令行参数
    import sys
    if len(sys.argv) > 1:
        command = sys.argv[1]
        if command == 'daily':
            scheduler.run_daily_report()
        elif command == 'weekly':
            scheduler.run_weekly_report()
        else:
            print(f"Unknown command: {command}")
            print("Usage: python auto_report.py [daily|weekly]")
    else:
        # 默认生成日报
        scheduler.run_daily_report()


if __name__ == '__main__':
    main()
```

---

## 配置文件示例

```json
{
  "db_host": "localhost",
  "db_port": 3306,
  "db_user": "report_user",
  "db_password": "your_password",
  "db_name": "restaurant_db",
  
  "smtp_host": "smtp.qiye.aliyun.com",
  "smtp_port": 587,
  "smtp_user": "report@yourcompany.com",
  "smtp_password": "email_password",
  "email_recipients": ["manager@yourcompany.com", "boss@yourcompany.com"],
  
  "dingtalk_webhook": "https://oapi.dingtalk.com/robot/send?access_token=xxx",
  "dingtalk_secret": "your_secret",
  
  "output_dir": "./reports"
}
```

---

## 定时任务配置

### Linux Crontab
```bash
# 每天上午9点生成日报
0 9 * * * cd /path/to/report && python auto_report.py daily

# 每周一上午9点生成周报
0 9 * * 1 cd /path/to/report && python auto_report.py weekly
```

### Windows 任务计划程序
```powershell
# 创建日报任务
schtasks /create /tn "DailyReport" /tr "python C:\reports\auto_report.py daily" /sc daily /st 09:00

# 创建周报任务
schtasks /create /tn "WeeklyReport" /tr "python C:\reports\auto_report.py weekly" /sc weekly /d MON /st 09:00
```

---

## 报表输出示例

Excel报表包含以下工作表：

1. **数据概览** - 关键指标卡片、品类TOP10、热销商品TOP15
2. **门店对比** - 各门店销售数据对比
3. **时段分析** - 24小时销售分布
4. **原始数据** - 完整明细数据

图表输出：
- 销售趋势图
- 品类占比饼图
- 时段分布柱状图
